import os
import uuid
from abc import ABC, abstractmethod
from itertools import islice, groupby
from pathlib import Path
from urllib.parse import quote

from django.http import HttpResponse
from openpyxl.reader.excel import load_workbook

from config.settings import BASE_DIR
from hospital.domain.valueobject.export_report import (
    BillingListRow,
    VotingManagementListRow,
)
from hospital.models import ElectionLedger


class AbstractExport(ABC):
    def __init__(self, temp_folder: Path, chunk_size: int, start_row: int):
        self.temp_folder = temp_folder
        self.chunk_size = chunk_size
        self.start_row = start_row
        os.makedirs(self.temp_folder, exist_ok=True)

    def create_unique_filename(self):
        return str(self.temp_folder / f"{str(uuid.uuid4())}.xlsx")

    def get_excel_data(self, filename: str) -> bytes:
        with open(self.temp_folder / filename, "rb") as f:
            return f.read()

    @abstractmethod
    def export(self, *args, **kwargs):
        pass


class ExportBillingService(AbstractExport):
    def __init__(self, temp_folder: Path):
        super().__init__(temp_folder, chunk_size=15, start_row=5)

    def export(self, election_id) -> HttpResponse:
        ledgers = (
            ElectionLedger.objects.filter(election_id=election_id)
            .select_related("vote_ward")
            .order_by("vote_ward__name")
        )

        wb = load_workbook(BASE_DIR / "hospital/domain/service/xlsx/billing_list.xlsx")
        filename = self.create_unique_filename()

        for ward_name, group in groupby(ledgers, key=lambda x: x.vote_ward.name):
            sheet_counter = 1
            ledgers_iter = iter(group)
            while True:
                chunk = list(islice(ledgers_iter, self.chunk_size))
                if not chunk:
                    break

                new_worksheet = wb.copy_worksheet(wb["ひな形"])
                new_worksheet.title = (
                    ward_name
                    if sheet_counter == 1
                    else f"{ward_name} ({sheet_counter})"
                )
                sheet_counter += 1

                for i, ledger in enumerate(chunk, start=0):
                    row = BillingListRow(ledger)
                    current_row_index = self.start_row + i
                    new_worksheet.cell(
                        row=current_row_index, column=1, value=row.address
                    )  # A列
                    new_worksheet.cell(
                        row=current_row_index, column=2, value=row.voter_name
                    )  # B列
                    new_worksheet.cell(
                        row=current_row_index, column=3, value=row.date_of_birth
                    )  # C列
        del wb["ひな形"]
        wb.save(filename)

        try:
            excel_data = self.get_excel_data(filename)
            response = HttpResponse(
                excel_data,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            header_content = f"attachment; filename*=UTF-8''{quote('請求者名簿.xlsx')}"
            response["Content-Disposition"] = header_content
        finally:
            os.remove(self.temp_folder / filename)

        return response


class VotingManagementService(AbstractExport):
    def __init__(self, temp_folder: Path):
        super().__init__(temp_folder, chunk_size=11, start_row=13)

    def export(self, election_id) -> HttpResponse:
        # TODO: 元帳ではなく投票管理に関連するデータを取得します（出力済みのものは印刷しない、モードつける？）
        ledgers = (
            ElectionLedger.objects.filter(election_id=election_id)
            .select_related("vote_ward")
            .order_by("vote_ward__name")
        )

        wb = load_workbook(
            BASE_DIR / "hospital/domain/service/xlsx/voting_management_list.xlsx"
        )
        filename = self.create_unique_filename()

        for ward_name, group in groupby(ledgers, key=lambda x: x.vote_ward.name):
            sheet_counter = 1
            ledgers_iter = iter(group)
            while True:
                chunk = list(islice(ledgers_iter, self.chunk_size))
                if not chunk:
                    break

                new_worksheet = wb.copy_worksheet(wb["ひな形"])
                new_worksheet.title = (
                    ward_name
                    if sheet_counter == 1
                    else f"{ward_name} ({sheet_counter})"
                )
                sheet_counter += 1

                for i, ledger in enumerate(chunk, start=0):
                    row = VotingManagementListRow(ledger)
                    current_row_index = self.start_row + i * 2
                    new_worksheet.cell(
                        row=current_row_index, column=2, value=str(i + 1)
                    )  # B列
                    new_worksheet.cell(
                        row=current_row_index, column=3, value=row.address
                    )  # C列
                    new_worksheet.cell(
                        row=current_row_index, column=4, value=row.voter_name
                    )  # D列
                    new_worksheet.cell(
                        row=current_row_index, column=5, value=row.billing_method
                    )  # E列
                    new_worksheet.cell(
                        row=current_row_index,
                        column=6,
                        value=row.proxy_billing_request_date,
                    )  # F列
                    new_worksheet.cell(
                        row=current_row_index, column=7, value=row.proxy_billing_date
                    )  # G列
                    new_worksheet.cell(
                        row=current_row_index, column=8, value=row.ballot_received_date
                    )  # H列
                    new_worksheet.cell(
                        row=current_row_index, column=9, value=row.vote_date
                    )  # I列
                    new_worksheet.cell(
                        row=current_row_index, column=10, value=row.vote_place
                    )  # J列
                    new_worksheet.cell(
                        row=current_row_index, column=11, value=row.voter_witness
                    )  # K列
                    new_worksheet.cell(
                        row=current_row_index,
                        column=12,
                        value=row.applied_for_proxy_voting,
                    )  # L列
                    new_worksheet.cell(
                        row=current_row_index, column=15, value=row.delivery_date
                    )  # O列
        del wb["ひな形"]
        wb.save(filename)

        try:
            excel_data = self.get_excel_data(filename)
            response = HttpResponse(
                excel_data,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            header_content = (
                f"attachment; filename*=UTF-8''{quote('不在者投票事務処理簿.xlsx')}"
            )
            response["Content-Disposition"] = header_content
        finally:
            os.remove(self.temp_folder / filename)

        return response
