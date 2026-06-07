import dataclasses
import glob
import io
import os
import re
import shutil
import zipfile
from pathlib import Path

from django.contrib import messages
from django.core.files.uploadedfile import UploadedFile
from django.core.management import call_command
from django.db.models import Prefetch, Sum, Count
from django.http import (
    HttpResponse,
    HttpResponseRedirect,
    JsonResponse,
    Http404,
)
from django.shortcuts import redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import (
    ListView,
    CreateView,
    DetailView,
    TemplateView,
    FormView,
)
from openpyxl import Workbook, load_workbook

from lib.geo.valueobject.coord import XarvioCoord
from lib.zipfileservice import ZipFileService
from soil_analysis.domain.repository.chemical_import_error import (
    ChemicalImportErrorRepository,
)
from soil_analysis.domain.repository.chemical_measurement import (
    SoilChemicalMeasurementRepository,
)
from soil_analysis.domain.repository.company import CompanyRepository
from soil_analysis.domain.repository.hardness_import_error import (
    HardnessImportErrorRepository,
)
from soil_analysis.domain.repository.hardness_measurement import (
    SoilHardnessMeasurementRepository,
)
from soil_analysis.domain.repository.land import LandRepository
from soil_analysis.domain.repository.land_block import LandBlockRepository
from soil_analysis.domain.repository.land_ledger import LandLedgerRepository
from soil_analysis.domain.repository.land_review import LandReviewRepository
from soil_analysis.domain.service.chemical_import_service import (
    ChemicalImportService,
)
from soil_analysis.domain.service.hardness_import_service import (
    HardnessImportService,
)
from soil_analysis.domain.service.hardness_measurement_service import (
    HardnessMeasurementService,
)
from soil_analysis.domain.service.hardness_plot_generation import (
    HardnessPlotGenerationService,
)
from soil_analysis.domain.service.kml import KmlService
from soil_analysis.domain.service.photo_processing import PhotoProcessingService
from soil_analysis.domain.valueobject.photo_processing.photo_spot import PhotoSpot
from soil_analysis.domain.valueobject.report.chemical_assessment import (
    ChemicalAssessmentVO,
)
from soil_analysis.domain.valueobject.report.fields import REPORT_FIELDS
from soil_analysis.domain.valueobject.report.hardness_assessment import (
    HardnessBlockAssessment,
)
from soil_analysis.forms import (
    CompanyCreateForm,
    LandCreateForm,
    UploadForm,
    ChemicalUploadForm,
    CsvGenerateForm,
    LandLedgerCreateForm,
)
from soil_analysis.models import (
    Company,
    Land,
    SoilChemicalMeasurement,
    LandLedger,
    SoilHardnessMeasurement,
    RouteSuggestImport,
    RokunoheLandRegistry,
    JmaCity,
    JmaRegion,
    JmaPrefecture,
    JmaWeather,
    JmaWarning,
)

SAMPLING_TIMES_PER_BLOCK = 5


class Home(ListView):
    model = Company
    template_name = "soil_analysis/home.html"
    context_object_name = "companies"

    def get_queryset(self):
        return (
            Company.objects.filter(category__name="農業法人")
            .prefetch_related("land_set")
            .order_by("name")
        )


class CompanyCreateView(CreateView):
    model = Company
    template_name = "soil_analysis/company/create.html"
    form_class = CompanyCreateForm

    def get_success_url(self):
        return reverse("soil:company_detail", kwargs={"pk": self.object.pk})


class CompanyDetailView(DetailView):
    model = Company
    template_name = "soil_analysis/company/detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company_id = self.kwargs["pk"]
        if not company_id:
            raise Http404("Company ID is required.")

        try:
            company = CompanyRepository.get_company_by_id(company_id)
        except Company.DoesNotExist:
            raise Http404("Company does not exist.")

        company_lands = list(Land.objects.filter(company_id=company_id))
        context["land_ledger_map"] = LandRepository.get_land_to_ledgers_map(
            company_lands
        )
        context["company"] = company

        return context


class LandCreateView(CreateView):
    model = Land
    template_name = "soil_analysis/land/create.html"
    form_class = LandCreateForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["company"] = Company(pk=self.kwargs["company_id"])
        return context

    def get_initial(self):
        initial = super().get_initial()
        # URLパラメータから圃場名を事前設定
        suggested_name = self.request.GET.get("suggested_name")
        if suggested_name:
            initial["name"] = suggested_name
        return initial

    def form_valid(self, form):
        form.instance.company_id = self.kwargs["company_id"]
        if form.instance.center:
            # center フィールドの値を正規化（スペースを除去）
            lat, lon = form.instance.center.split(",")
            form.instance.center = ",".join([lat.strip(), lon.strip()])
        return super().form_valid(form)

    def get_success_url(self):
        # suggested_nameがある場合（硬度測定データからの圃場作成）は硬度測定画面に戻る
        suggested_name = self.request.GET.get("suggested_name")
        if suggested_name:
            return reverse("soil:hardness_success")

        # 通常の圃場作成の場合は圃場詳細画面に遷移
        return reverse(
            "soil:land_detail",
            kwargs={"company_id": self.kwargs["company_id"], "pk": self.object.pk},
        )


class PrefecturesView(View):

    @staticmethod
    def get(request):
        prefectures = JmaPrefecture.objects.all()
        data = {"prefectures": list(prefectures.values("id", "name"))}
        return JsonResponse(data)


class PrefectureCitiesView(View):
    """
    圃場新規作成時のフォームで prefecture がonChangeした際に非同期で、該当するcityを取得
    """

    @staticmethod
    def get(request, prefecture_id):
        regions = JmaRegion.objects.filter(jma_prefecture__id=prefecture_id)
        cities = JmaCity.objects.filter(jma_region__in=regions)

        data = {"cities": list(cities.values("id", "name"))}
        return JsonResponse(data)


class LandDetailView(DetailView):
    model = Land
    template_name = "soil_analysis/land/detail.html"

    def get_queryset(self):
        weather_prefetch = Prefetch(
            "jma_city__jma_region__jmaweather_set",
            queryset=JmaWeather.objects.all(),
            to_attr="weathers",
        )
        warning_prefetch = Prefetch(
            "jma_city__jma_region__jmawarning_set",
            queryset=JmaWarning.objects.all(),
            to_attr="warnings",
        )
        land_ledger_prefetch = Prefetch(
            "landledger_set",
            queryset=LandLedger.objects.select_related("land_period", "crop"),
            to_attr="ledgers",
        )
        return (
            super()
            .get_queryset()
            .prefetch_related(weather_prefetch, warning_prefetch, land_ledger_prefetch)
        )


class StandardReportView(ListView):
    """
    化学分析レポート（通知表）の一覧表示

    Attributes:
        model: 使用するモデル
        template_name: 使用するテンプレート名
    """

    model = SoilChemicalMeasurement
    template_name = "soil_analysis/land_report/standard_report.html"

    def get_queryset(self):
        """
        特定の帳簿に紐づく化学分析データを取得する。

        Returns:
            SoilChemicalMeasurement のクエリセット
        """
        land_ledger_id = self.kwargs["land_ledger_id"]
        return super().get_queryset().filter(land_ledger_id=land_ledger_id)

    def get_context_data(self, **kwargs):
        """
        レポート表示に必要なコンテキストデータを生成する。

        Args:
            **kwargs: 追加のコンテキスト引数

        Returns:
            テンプレートに渡すコンテキスト辞書
        """
        context = super().get_context_data(**kwargs)
        land_ledger_id = self.kwargs["land_ledger_id"]

        try:
            land_ledger = LandLedgerRepository.get_with_details(land_ledger_id)
            company = land_ledger.land.company
        except LandLedger.DoesNotExist:
            raise Http404("Land Ledger does not exist.")

        context["company"] = company
        context["land"] = land_ledger.land
        context["land_ledger"] = land_ledger

        # 圃場単位の化学分析データを取得
        context["soil_analysis"] = SoilChemicalMeasurementRepository.get_by_ledger(
            land_ledger
        )

        context["land_scores"] = []
        context["land_review"] = LandReviewRepository.get_by_ledger(land_ledger)
        context["land_scores_dict"] = {}
        context["land_blocks"] = LandBlockRepository.get_all()
        context["report_fields"] = REPORT_FIELDS

        # 化学判定VOの生成
        context["chemical_assessment"] = ChemicalAssessmentVO.from_measurements(
            [context["soil_analysis"]] if context["soil_analysis"] else []
        )

        # 硬度判定VOの生成
        context["hardness_assessment"] = (
            HardnessMeasurementService.get_hardness_assessment(land_ledger)
        )

        context["hardness_thresholds"] = {
            "low": HardnessBlockAssessment.THRESHOLD_LOW,
            "high": HardnessBlockAssessment.THRESHOLD_HIGH,
            "max": HardnessBlockAssessment.MAX_SCALE,
        }

        # グリッドの順序:
        # C3, B3, A3
        # C2, B2, A2
        # C1, B1, A1
        blocks = []
        for row in ["3", "2", "1"]:
            for col in ["C", "B", "A"]:
                name = f"{col}{row}"
                block = next(
                    (b for b in context["land_blocks"] if b.name == name), None
                )
                if block:
                    blocks.append(block)
        context["ordered_blocks"] = blocks

        # ブロックと評価データを紐付け
        context["ordered_blocks_with_assessment"] = [
            {"block": b, "assessment": context["hardness_assessment"].get_block(b.name)}
            for b in blocks
        ]

        return context


class LandLedgerCreateAjaxView(View):
    """
    帳簿（LandLedger）の新規作成Ajax処理
    Field Group画面から直接新規帳簿を作成する際に使用
    """

    @staticmethod
    def post(request):
        form = LandLedgerCreateForm(request.POST)

        if form.is_valid():
            try:
                # 新規帳簿を作成
                land_ledger = form.save()

                # 成功レスポンス
                response_data = {
                    "success": True,
                    "message": f"帳簿「{land_ledger.land.name} - {land_ledger.land_period.year} {land_ledger.land_period.name}」を作成しました。",
                    "land_ledger": {
                        "id": land_ledger.id,
                        "display_name": f"{land_ledger.land.company.name} - {land_ledger.land.name} ({land_ledger.land_period.year} {land_ledger.land_period.name})",
                    },
                }

                return JsonResponse(response_data)

            except Exception as e:
                return JsonResponse(
                    {
                        "success": False,
                        "message": f"帳簿の作成中にエラーが発生しました: {str(e)}",
                    },
                    status=500,
                )

        else:
            # フォームバリデーションエラー
            errors = []
            for field, error_list in form.errors.items():
                field_label = form.fields[field].label or field
                for error in error_list:
                    errors.append(f"{field_label}: {error}")

            return JsonResponse(
                {"success": False, "message": "エラー：\n" + "\n".join(errors)},
                status=400,
            )

    def get(self, request):
        """
        フォーム表示用データをAjaxで取得
        フォルダ名から圃場を推定する処理も含む
        """
        folder_name = request.GET.get("folder_name", "")

        # フォーム初期化
        form = LandLedgerCreateForm()

        response_data = LandLedgerRepository.get_form_data_for_ajax(folder_name)
        response_data["form_html"] = self._render_form_fields(form)

        return JsonResponse(response_data)

    @staticmethod
    def _render_form_fields(form):
        """フォームフィールドのHTMLを生成（簡易版）"""
        return {
            field_name: {
                "widget_type": field.widget.__class__.__name__,
                "required": field.required,
                "label": field.label,
                "help_text": field.help_text,
            }
            for field_name, field in form.fields.items()
        }


class HardnessUploadView(FormView):
    template_name = "soil_analysis/hardness/form.html"
    form_class = UploadForm
    success_url = reverse_lazy("soil:hardness_success")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["csv_generate_form"] = CsvGenerateForm()
        return context

    def form_valid(self, form):
        # Zipを処理してバッチ実行
        app_name = self.request.resolver_match.app_name
        upload_folder = ZipFileService.handle_uploaded_zip(
            self.request.FILES["file"], app_name
        )
        if os.path.exists(upload_folder):
            # エラーログのクリア
            HardnessImportErrorRepository.delete_all()

            csv_files = glob.glob(
                os.path.join(upload_folder, "**/*.csv"), recursive=True
            )
            for csv_file in csv_files:
                parent_folder = os.path.basename(os.path.dirname(csv_file))
                file_name = os.path.basename(csv_file)

                parse_result = HardnessImportService.parse_csv(csv_file)

                if parse_result.errors:
                    for error_msg in parse_result.errors:
                        HardnessImportErrorRepository.create(
                            file=file_name,
                            folder=parent_folder,
                            message=error_msg,
                        )
                    continue

                HardnessImportService.save_import_data(parse_result.rows)

            try:
                shutil.rmtree(upload_folder)
            except (PermissionError, OSError):
                # ファイル削除エラーは無視して続行
                # OneDriveなどの同期フォルダではこの例外が発生することがある
                pass

        return super().form_valid(form)


class ChemicalUploadView(FormView):
    template_name = "soil_analysis/chemical/form.html"
    form_class = ChemicalUploadForm

    def form_valid(self, form):
        upload_file = self.request.FILES["file"]
        if not upload_file.name.lower().endswith(".xlsx"):
            messages.error(self.request, "xlsxファイルを指定してください。")
            return self.form_invalid(form)

        try:
            workbook = load_workbook(upload_file, data_only=True)
            if len(workbook.sheetnames) != 1:
                messages.error(
                    self.request, "Excelファイルのシート数は1枚にしてください。"
                )
                return self.form_invalid(form)

            worksheet = workbook.active
            parse_result = ChemicalImportService.parse_kawada_worksheet(worksheet)

            if parse_result.errors:
                for error in parse_result.errors:
                    messages.error(self.request, error)
                return self.form_invalid(form)

            if not parse_result.rows:
                messages.error(self.request, "取り込み対象行がありません。")
                return self.form_invalid(form)

            # セッションに保存
            rows_data = []
            for row in parse_result.rows:
                rows_data.append(
                    {
                        "row_data": dataclasses.asdict(row),
                        "selected_ledger_id": None,
                        "status": "pending",
                    }
                )

            self.request.session["chemical_import_session"] = {
                "rows": rows_data,
                "total_rows": len(rows_data),
                "source_file": upload_file.name,
            }

            return redirect("soil:chemical_success")

        except Exception as e:
            messages.error(self.request, f"取り込み中にエラーが発生しました: {str(e)}")
            return self.form_invalid(form)


class ChemicalAssociationView(TemplateView):
    template_name = "soil_analysis/chemical/association/list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        import_session = self.request.session.get("chemical_import_session")
        if not import_session:
            return context

        rows = import_session.get("rows", [])
        confirmed_count = sum(1 for r in rows if r["status"] == "confirmed")

        # 圃場名から候補帳簿を一括取得
        land_names = [row["row_data"]["land_name"] for row in rows]
        suggested_map = ChemicalImportService.get_suggested_ledgers_for_names(
            land_names, base_ledger_id=import_session.get("base_ledger_id")
        )

        # 選択済み帳簿を一括取得してキャッシュ
        selected_ledger_ids = [
            row["selected_ledger_id"] for row in rows if row["selected_ledger_id"]
        ]
        selected_ledgers_map = {}
        if selected_ledger_ids:
            selected_ledgers_map = {
                l.id: l
                for l in LandLedger.objects.select_related(
                    "land", "land__company", "land_period"
                ).filter(id__in=selected_ledger_ids)
            }

        # 表示用に各行の候補帳簿などを付与（セッションは汚さない）
        display_rows = []
        for i, row in enumerate(rows):
            land_name = row["row_data"]["land_name"]
            suggested_ledgers = suggested_map.get(land_name, [])

            # 選択済み帳簿の情報を取得
            selected_ledger = selected_ledgers_map.get(row["selected_ledger_id"])

            display_rows.append(
                {
                    "index": i,
                    "row_number": row["row_data"]["row_number"],
                    "land_name": land_name,
                    "status": row["status"],
                    "suggested_count": len(suggested_ledgers),
                    "selected_ledger": selected_ledger,
                }
            )

        context.update(
            {
                "rows": display_rows,
                "total_rows": import_session.get("total_rows"),
                "confirmed_count": confirmed_count,
                "all_confirmed": confirmed_count == import_session.get("total_rows"),
            }
        )
        return context

    @staticmethod
    def post(request, *args, **kwargs):
        if "btn_save_all" in request.POST:
            import_session = request.session.get("chemical_import_session")
            if not import_session:
                messages.error(
                    request, "セッションが期限切れです。最初からやり直してください。"
                )
                return redirect("soil:chemical_upload")

            rows = import_session.get("rows", [])
            if any(r["status"] != "confirmed" for r in rows):
                messages.error(request, "未確定の行があります。")
                return redirect("soil:chemical_association")

            # 保存実行
            save_data = []
            for r in rows:
                save_data.append(
                    {
                        "row_data": r["row_data"],
                        "land_ledger_id": r["selected_ledger_id"],
                    }
                )

            try:
                source_file = import_session.get("source_file")
                result = ChemicalImportService.save_import_data(
                    save_data, source_file=source_file
                )
                request.session["chemical_import_result"] = result
                # セッションクリア
                del request.session["chemical_import_session"]
                return redirect("soil:chemical_association_success")
            except Exception as e:
                messages.error(request, f"保存中にエラーが発生しました: {str(e)}")
                return redirect("soil:chemical_association")

        return redirect("soil:chemical_association")


class ChemicalUploadSuccessView(TemplateView):
    template_name = "soil_analysis/chemical/success.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        import_session = self.request.session.get("chemical_import_session")
        if not import_session:
            return context

        rows = import_session.get("rows", [])
        context["total_records"] = len(rows)

        summary_map = {}
        for row in rows:
            land_name = row["row_data"].get("land_name") or ""
            if land_name not in summary_map:
                summary_map[land_name] = 0
            summary_map[land_name] += 1

        context["row_summary"] = [
            {"land_name": key, "count": value} for key, value in summary_map.items()
        ]
        return context


class ChemicalAssociationRowView(TemplateView):
    template_name = "soil_analysis/chemical/association/field_row.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        row_index = self.kwargs.get("row_index")
        import_session = self.request.session.get("chemical_import_session")

        if not import_session or row_index >= len(import_session["rows"]):
            raise Http404("Invalid row index")

        row = import_session["rows"][row_index]
        land_name = row["row_data"]["land_name"]

        suggested_ledgers = ChemicalImportService.get_suggested_ledgers(
            land_name, base_ledger_id=import_session.get("base_ledger_id")
        )
        all_ledgers = LandLedger.objects.select_related(
            "land", "land__company", "land_period"
        ).order_by("-id")[:200]
        suggested_ids = {ledger.id for ledger in suggested_ledgers}
        fallback_ledgers = [
            ledger for ledger in all_ledgers if ledger.id not in suggested_ids
        ]

        context.update(
            {
                "row_index": row_index,
                "row_data": row["row_data"],
                "selected_ledger_id": row["selected_ledger_id"],
                "suggested_ledgers": suggested_ledgers,
                "fallback_ledgers": fallback_ledgers,
                "total_rows": import_session.get("total_rows"),
            }
        )
        return context

    def post(self, request, *args, **kwargs):
        row_index = self.kwargs.get("row_index")
        import_session = request.session.get("chemical_import_session")

        if not import_session or row_index >= len(import_session["rows"]):
            return redirect("soil:chemical_association")

        ledger_id = request.POST.get("land_ledger")
        if ledger_id:
            import_session["rows"][row_index]["selected_ledger_id"] = int(ledger_id)
            import_session["rows"][row_index]["status"] = "confirmed"
            request.session.modified = True

            # 次の未確定行へ
            next_row_index = row_index + 1
            if next_row_index < len(import_session["rows"]):
                return redirect(
                    "soil:chemical_association_field_row", row_index=next_row_index
                )
            else:
                return redirect("soil:chemical_association")

        return redirect("soil:chemical_association_field_row", row_index=row_index)


class ChemicalSuccessView(TemplateView):
    template_name = "soil_analysis/chemical/association/success.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        result = self.request.session.get("chemical_import_result")
        context["import_errors"] = ChemicalImportErrorRepository.get_all()
        if result:
            context["created_count"] = result.get("created", 0)
            context["updated_count"] = result.get("updated", 0)
            context["total_count"] = context["created_count"] + context["updated_count"]
            context["display_record_count"] = context["total_count"]
            ledger_ids = result.get("ledger_ids", [])
            ledger_summary = []
            if ledger_ids:
                count_by_ledger = {
                    row["land_ledger_id"]: row["total"]
                    for row in SoilChemicalMeasurement.objects.filter(
                        land_ledger_id__in=ledger_ids
                    )
                    .values("land_ledger_id")
                    .annotate(total=Count("id"))
                }

                ledgers = (
                    LandLedger.objects.select_related(
                        "land", "land__company", "land_period"
                    )
                    .prefetch_related(
                        Prefetch(
                            "soil_chemical_measurement",
                            queryset=SoilChemicalMeasurement.objects.all(),
                            to_attr="measurement",
                        )
                    )
                    .filter(id__in=ledger_ids)
                )

                for ledger in ledgers:
                    ledger_summary.append(
                        {
                            "company_name": ledger.land.company.name,
                            "land_name": ledger.land.name,
                            "period_name": ledger.land_period.name,
                            "total": count_by_ledger.get(ledger.id, 0),
                        }
                    )
            context["ledger_summary"] = ledger_summary
            context["import_datetime"] = timezone.now()
        return context


class HardnessDeleteAllView(View):
    """
    SoilHardnessMeasurementテーブルの全データ削除
    開発・テスト環境での使用を想定
    """

    @staticmethod
    def post(request, *args, **kwargs):
        try:
            # 削除前のレコード数を取得
            count = SoilHardnessMeasurement.objects.count()

            # 全データ削除
            SoilHardnessMeasurement.objects.all().delete()

            messages.success(
                request,
                f"SoilHardnessMeasurementテーブルの全データ（{count}件）を削除しました。",
            )
        except Exception as e:
            messages.error(request, f"削除中にエラーが発生しました: {str(e)}")

        return HttpResponseRedirect(reverse("soil:hardness_upload"))


class ChemicalDownloadSampleView(View):
    """
    連続登録テスト用のサンプルExcel一式をZIPでダウンロード提供
    """

    @staticmethod
    def get(request, *args, **kwargs):
        zip_buffer = io.BytesIO()
        sample_files = {
            "chemical_stage01.xlsx": ChemicalDownloadSampleView._create_workbook_bytes(
                stage=1
            ),
            "chemical_stage02.xlsx": ChemicalDownloadSampleView._create_workbook_bytes(
                stage=2
            ),
            "chemical_duplicate.xlsx": ChemicalDownloadSampleView._create_workbook_bytes(
                stage=1
            ),
            "README.txt": ChemicalDownloadSampleView._create_readme().encode("utf-8"),
        }

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for filename, content in sample_files.items():
                zip_file.writestr(filename, content)

        response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = 'attachment; filename="chemical_samples.zip"'
        return response

    @staticmethod
    def _create_workbook_bytes(stage: int) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "川田研究所レポート"

        worksheet.append(
            [
                "Agsoil株式会社",
                None,
                "御中",
                None,
                None,
                None,
                f"2026.6.{stage}",
            ]
        )
        worksheet.append([])
        worksheet.append(
            [
                "分析番号",
                "氏名",
                "圃場名",
                "栽培作物",
                "EC (mS/cm)",
                "pH",
                "meq/100g",
                "CaO (mg/100g)",
                "MgO (mg/100g)",
                "K2O (mg/100g)",
                "石灰飽和度 %",
                "苦土飽和度 %",
                "加里飽和度 %",
                "塩基飽和度 %",
                "P2O5 (mg/100g)",
                "リン吸 (mg/100g)",
                "NH4-N (mg/100g)",
                "NO3-N (mg/100g)",
                "腐植 %",
                "仮比重",
            ]
        )

        base_analysis_number = 2606000 + (stage * 100)
        land_names = (
            "FIELD001（点検用圃場）",
            "FIELD002（点検用圃場）",
            "FIELD003（点検用圃場）",
        )
        for row_index, land_name in enumerate(land_names, start=1):
            worksheet.append(
                ChemicalDownloadSampleView._create_chemical_row(
                    analysis_number=base_analysis_number + row_index,
                    land_name=land_name,
                    crop_name=("レタス", "キャベツ", "トマト")[row_index - 1],
                    value_offset=(stage * 10) + row_index,
                )
            )

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _create_chemical_row(
        analysis_number: int, land_name: str, crop_name: str, value_offset: int
    ) -> list:
        return [
            str(analysis_number),
            "テスト担当",
            land_name,
            crop_name,
            round(0.20 + value_offset * 0.01, 3),
            round(6.10 + value_offset * 0.02, 2),
            round(11.0 + value_offset * 0.3, 1),
            round(300.0 + value_offset * 3.0, 1),
            round(45.0 + value_offset * 1.2, 1),
            round(22.0 + value_offset * 0.8, 1),
            round(85.0 + value_offset * 0.7, 1),
            round(18.0 + value_offset * 0.4, 1),
            round(4.0 + value_offset * 0.2, 1),
            round(105.0 + value_offset * 0.9, 1),
            round(600.0 + value_offset * 4.0, 1),
            round(520.0 + value_offset * 5.0, 1),
            round(0.5 + value_offset * 0.03, 2),
            round(1.0 + value_offset * 0.08, 2),
            round(2.0 + value_offset * 0.06, 2),
            round(1.00 + value_offset * 0.01, 2),
        ]

    @staticmethod
    def _create_readme() -> str:
        return "\n".join(
            [
                "化学分析 連続アップロード検証用データ",
                "",
                "1. chemical_stage01.xlsx を soil:chemical_upload へ投入してください。",
                "2. chemical_stage02.xlsx を続けて投入してください。",
                "3. chemical_duplicate.xlsx は stage01 と同じ圃場名を含むため、同じ台帳に割り当てた場合の更新検証に使います。",
                "",
                "各Excelは川田研究所形式の列順に合わせています。",
            ]
        )


class ChemicalDeleteAllView(View):
    """
    SoilChemicalMeasurementテーブルの全データ削除
    開発・テスト環境での使用を想定
    """

    @staticmethod
    def post(request, *args, **kwargs):
        try:
            # 削除前のレコード数を取得
            count = SoilChemicalMeasurement.objects.count()

            # 全データ削除
            SoilChemicalMeasurement.objects.all().delete()

            messages.success(
                request,
                f"SoilChemicalMeasurementテーブルの全データ（{count}件）を削除しました。",
            )
        except Exception as e:
            messages.error(request, f"削除中にエラーが発生しました: {str(e)}")

        return HttpResponseRedirect(reverse("soil:chemical_upload"))


class HardnessGenerateDummyCsvView(View):
    """
    テスト用CSVを生成してZIPファイルでダウンロード提供
    """

    @staticmethod
    def post(request, *args, **kwargs):
        try:
            want_to_create_dataset_round = int(
                request.POST.get("want_to_create_dataset_round", 2)
            )
            num_fields = int(request.POST.get("num_fields", 3))

            # CSVを生成して出力パスを取得
            csv_output_path = call_command(
                "hardness_generate_dummy_csv",
                f"--num_fields={num_fields}",
                f"--want_to_create_dataset_round={want_to_create_dataset_round}",
            )

            if csv_output_path and os.path.exists(csv_output_path):
                # ZIP化してダウンロード
                response = ZipFileService.create_zip_download(
                    csv_output_path, "hardness_samples.zip"
                )

                # 一時ディレクトリを削除
                shutil.rmtree(Path(csv_output_path).parent)

                return response
            else:
                messages.error(request, "CSV生成に失敗しました。")
                return HttpResponseRedirect(reverse("soil:hardness_upload"))

        except Exception as e:
            messages.error(request, f"CSV生成中にエラーが発生しました: {str(e)}")
            return HttpResponseRedirect(reverse("soil:hardness_upload"))


class HardnessSuccessView(TemplateView):
    template_name = "soil_analysis/hardness/success.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["import_errors"] = HardnessImportErrorRepository.get_all()

        # CSVインポートされた硬度測定データをフォルダ名でグループ化し、各フォルダの統計情報を取得
        folder_stats = SoilHardnessMeasurementRepository.get_folder_stats(
            associated_only=False
        )

        # フォルダ名に基づいて新規登録が必要な圃場を特定
        missing_lands = []
        for stats in folder_stats:
            folder_name = stats.folder
            land_name = self._extract_land_name_from_folder(folder_name)

            if land_name and not LandRepository.exists_by_name(land_name):
                missing_lands.append(
                    {"folder_name": folder_name, "suggested_land_name": land_name}
                )

        context["folder_stats"] = folder_stats
        context["total_records"] = len(
            SoilHardnessMeasurementRepository.get_folder_stats()
        )
        context["missing_lands"] = missing_lands

        # 圃場作成用の会社一覧を追加（農業法人のみ）
        if missing_lands:
            context["companies"] = CompanyRepository.get_by_category(1)

        return context

    @staticmethod
    def _extract_land_name_from_folder(folder_name: str) -> str:
        """
        フォルダ名から圃場名を抽出する。

        Args:
            folder_name: 変換対象のフォルダ名

        Returns:
            抽出された圃場名

        変換パターン例:
        - "静岡ススムA1_20230701" → "静岡ススムA1"
        - "静岡ススムA1_20230701_1" → "静岡ススムA1"
        - "静岡ススムA120230701" → "静岡ススムA1"
        - "静岡ススムA120230701_extra" → "静岡ススムA1"
        - "静岡ススムA1" → "静岡ススムA1"
        """
        # アンダースコアがある場合は最初の部分を取得
        if "_" in folder_name:
            base_name = folder_name.split("_")[0]
        else:
            base_name = folder_name

        # 8桁日付パターン（YYYYMMDD）以降をすべて除去
        date_pattern = r"\d{8}.*$"
        base_name = re.sub(date_pattern, "", base_name)

        return base_name.strip()


class HardnessAssociationView(ListView):
    model = SoilHardnessMeasurement
    template_name = "soil_analysis/hardness/association/list.html"

    def get_queryset(self, **kwargs):
        return HardnessImportService.get_folder_groups_for_association()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["land_ledgers"] = LandLedgerRepository.get_all()

        context["total_groups"] = HardnessImportService.get_total_groups_count()
        context["processed_groups"] = HardnessImportService.get_processed_groups_count()

        return context

    @staticmethod
    def post(request, **kwargs):
        """
        圃場グループ別処理：
        - 単一圃場グループの帳簿選択処理
        - 処理完了後に次の未処理圃場へ自動遷移
        """

        # 圃場グループ処理ボタンが押された場合（この圃場を処理ボタン）
        if "btn_process_group" in request.POST:
            memory_anchor_str = request.POST.get("btn_process_group")
            try:
                if not memory_anchor_str:
                    messages.error(request, "メモリーアンカーが指定されていません")
                    return HttpResponseRedirect(reverse("soil:hardness_association"))

                memory_anchor = int(memory_anchor_str)
                return HttpResponseRedirect(
                    reverse(
                        "soil:hardness_association_field_group",
                        kwargs={"memory_anchor": memory_anchor},
                    )
                )
            except (ValueError, TypeError) as e:
                # memory_anchor_strが未定義の可能性があるため、エラー変数を使用
                messages.error(
                    request,
                    f"無効なメモリーアンカーです: {memory_anchor_str if 'memory_anchor_str' in locals() else 'unknown'}",
                )
                return HttpResponseRedirect(reverse("soil:hardness_association"))

        return HttpResponseRedirect(reverse("soil:hardness_association"))


class HardnessAssociationFieldGroupView(ListView):
    """
    単一圃場グループの帳簿選択画面
    """

    model = SoilHardnessMeasurement
    template_name = "soil_analysis/hardness/association/field_group.html"

    def get_queryset(self, **kwargs):
        memory_anchor = self.kwargs.get("memory_anchor")
        # メモリー番号からフォルダを特定し、そのフォルダの全データを取得
        sample_measurement = (
            SoilHardnessMeasurementRepository.get_first_by_memory_anchor(memory_anchor)
        )

        if sample_measurement:
            folder_name = sample_measurement.folder
            return SoilHardnessMeasurementRepository.get_all_by_folder(folder_name)
        return []

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        memory_anchor = self.kwargs.get("memory_anchor")

        # 該当圃場グループのフォルダ名から適切な帳簿を絞り込み
        measurements = self.get_queryset()
        folder_name = measurements[0].folder if measurements else ""

        # メモリー番号の範囲を計算
        if measurements:
            memory_numbers = list(set(m.set_memory for m in measurements))
            min_memory = min(memory_numbers)
            max_memory = max(memory_numbers)
        else:
            min_memory = max_memory = memory_anchor

        # フォルダ名に基づいて適切な帳簿のみを表示
        suitable_ledgers = HardnessImportService.get_suitable_ledgers(folder_name)

        context.update(
            {
                "memory_anchor": memory_anchor,
                "min_memory": min_memory,
                "max_memory": max_memory,
                "folder_name": folder_name,
                "land_ledgers": suitable_ledgers,
                "total_groups": HardnessImportService.get_total_groups_count(),
                "processed_groups": HardnessImportService.get_processed_groups_count(),
            }
        )
        return context

    def post(self, request, **kwargs):
        """圃場グループの帳簿選択処理"""
        form_land_ledger_id = int(request.POST.get("land_ledger"))

        # 処理対象のフォルダを特定
        memory_anchor = self.kwargs.get("memory_anchor")
        sample_measurement = (
            SoilHardnessMeasurementRepository.get_first_by_memory_anchor(memory_anchor)
        )

        if not sample_measurement:
            messages.error(request, "処理対象のデータが見つかりません")
            return HttpResponseRedirect(reverse("soil:hardness_association"))

        folder_name = sample_measurement.folder

        success = HardnessImportService.associate_with_ledger(
            folder_name, form_land_ledger_id
        )

        if success:
            messages.success(
                request,
                f"フォルダ「{folder_name}」の処理が完了しました",
            )
        else:
            messages.error(request, "紐付け処理に失敗しました")

        # 処理完了後は常にリスト画面に戻る（シンプル化）
        return HttpResponseRedirect(reverse("soil:hardness_association"))


class HardnessAssociationSuccessView(TemplateView):
    template_name = "soil_analysis/hardness/association/success.html"

    @staticmethod
    def post(request, *args, **kwargs):
        if "btn_generate_plots" in request.POST:
            HardnessPlotGenerationService.generate_and_save_plots()
            return HttpResponseRedirect(reverse("soil:home"))

        return HttpResponseRedirect(request.path)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        repository_context = LandLedgerRepository.get_association_success_context()
        context.update(repository_context)

        return context


class RouteSuggestUploadView(FormView):
    template_name = "soil_analysis/route_suggest/form.html"
    form_class = UploadForm
    success_url = reverse_lazy("soil:route_suggest_ordering")

    def form_valid(self, form):
        """
        Notes: Directions API の地点を制限する
         可能であれば、クエリでのユーザー入力を最大 10 地点に制限します。10 を超える地点を含むリクエストは、課金レートが高くなります。
         https://developers.google.com/maps/optimization-guide?hl=ja#routes
        """
        upload_file: UploadedFile = self.request.FILES["file"]
        kml_raw = upload_file.read().decode("utf-8")
        kml_service = KmlService()
        land_location_list = kml_service.parse_kml(kml_raw)

        if len(land_location_list) < 2:
            messages.error(self.request, "少なくとも 2 つの場所を指定してください")
            return redirect(self.request.META.get("HTTP_REFERER"))

        if len(land_location_list) > 10:
            messages.error(
                self.request,
                "GoogleMapsAPIのレート上昇制約により 10 地点までしか計算できません",
            )
            return redirect(self.request.META.get("HTTP_REFERER"))

        entities = []
        for land_location in land_location_list:
            coordinates_str = land_location.center.to_google().to_str()
            entity = RouteSuggestImport.objects.create(
                name=land_location.name, coord=coordinates_str
            )
            entities.append(entity)
        RouteSuggestImport.objects.all().delete()
        RouteSuggestImport.objects.bulk_create(entities)

        return super().form_valid(form)


class RouteSuggestOrderingView(ListView):
    model = RouteSuggestImport
    template_name = "soil_analysis/route_suggest/ordering.html"

    def post(self, request, *args, **kwargs):
        order_data = self.request.POST.get("order_data")

        try:
            if order_data:
                order_ids = order_data.split(",")
                for order, order_id in enumerate(order_ids, start=1):
                    route_suggest = RouteSuggestImport.objects.get(pk=order_id)
                    route_suggest.ordering = order
                    route_suggest.save()

            messages.success(request, "Data updated successfully")
            return redirect(reverse("soil:route_suggest_success"))

        except RouteSuggestImport.DoesNotExist:
            messages.error(request, "Invalid order data provided.")
            return redirect(request.META.get("HTTP_REFERER"))


class RouteSuggestSuccessView(TemplateView):
    template_name = "soil_analysis/route_suggest/success.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        route_suggest_imports = RouteSuggestImport.objects.all().order_by("ordering")
        company_list = []
        land_list = []
        for route_suggest_import in route_suggest_imports:
            company_name, land_name = route_suggest_import.name.split(" - ")
            company_list.append(company_name)
            land_list.append({"name": land_name, "coord": route_suggest_import.coord})

        context["company_list"] = company_list
        context["land_list"] = land_list
        context["coord_list"] = list(land["coord"] for land in land_list)
        context["google_maps_fe_api_key"] = os.getenv("GOOGLE_MAPS_FE_API_KEY")

        return context


class AssociatePictureAndLandView(ListView):
    model = Land
    template_name = "soil_analysis/picture_land_associate/form.html"
    context_object_name = "land_list"
    success_url = reverse_lazy("soil:associate_picture_and_land_result")

    def get_queryset(self):
        return Land.objects.all().order_by("pk")

    @staticmethod
    def get_dummy_photo_spots() -> list[XarvioCoord]:
        """
        テスト用の撮影位置データを返します
        注: GPSが正確な撮影位置を取得できるようになったらここを置き換える
        """
        return [
            XarvioCoord(longitude=137.64905, latitude=34.74424),  # 静岡ススムA1用
            XarvioCoord(longitude=137.64921, latitude=34.744),  # 静岡ススムA2用
            XarvioCoord(longitude=137.64938, latitude=34.74374),  # 静岡ススムA3用
            XarvioCoord(longitude=137.6496, latitude=34.7434),  # 静岡ススムA4用
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["photo_spots"] = self.get_dummy_photo_spots()
        return context

    def post(self, request, *args, **kwargs):
        if "photo_spot" not in request.POST:
            return self.render_to_response(self.get_context_data())

        spot_index = int(request.POST["photo_spot"])
        photo_spots = self.get_dummy_photo_spots()

        photo_spot = PhotoSpot(photo_spots[spot_index])

        service = PhotoProcessingService()
        nearest_land = service.find_nearest_land(photo_spot, list(self.get_queryset()))

        # セッションに結果を保存
        self.request.session["nearest_land_id"] = nearest_land.id
        self.request.session["photo_spot_coord"] = (
            photo_spot.original_position.to_google().to_str()
        )

        return HttpResponseRedirect(self.success_url)


class AssociatePictureAndLandResultView(TemplateView):
    template_name = "soil_analysis/picture_land_associate/result.html"

    # Googleマップルートのパラメータを定数として定義
    GOOGLE_MAPS_PARAMS = {
        "DISPLAY_OPTIONS": "!3m1!4b1",  # マップの表示オプション
        "ROUTE_PARAMS": "!4m2!4m1",  # ルート関連パラメータ
        "TRANSPORT_MODES": {
            "CAR": "!3e0",  # 自動車での移動
            "PUBLIC_TRANSPORT": "!3e1",  # 公共交通機関での移動
            "WALKING": "!3e2",  # 徒歩での移動
            "BICYCLE": "!3e3",  # 自転車での移動
        },
    }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        photo_spot_coord = self.request.session.get("photo_spot_coord")
        nearest_land_id = self.request.session.get("nearest_land_id")

        if nearest_land_id and photo_spot_coord:
            # 圃場idから圃場データを取得
            land = LandRepository.find_land_by_id(nearest_land_id)

            # ルートURL作成（徒歩ルート指定）
            context["route_url"] = (
                f"https://www.google.com/maps/dir/{photo_spot_coord}/{land.to_google().to_str()}/"
                f"data={self.GOOGLE_MAPS_PARAMS['DISPLAY_OPTIONS']}"
                f"{self.GOOGLE_MAPS_PARAMS['ROUTE_PARAMS']}"
                f"{self.GOOGLE_MAPS_PARAMS['TRANSPORT_MODES']['WALKING']}"
            )

            context["nearest_land"] = {
                "name": land.name,
                "location": land.to_google().to_str(),
                "area": land.area,
                "owner": land.owner.username,
            }
            context["photo_spot_coord"] = photo_spot_coord

        return context


class RokunoheLandRegistryListView(ListView):
    model = RokunoheLandRegistry
    template_name = "soil_analysis/rokunohe_land_registry/list.html"
    context_object_name = "registries"

    def get_queryset(self):
        return RokunoheLandRegistry.objects.all().order_by("id")

    @staticmethod
    def _to_chobu(area_m2: int) -> str:
        one_tan = 991.736
        one_se = one_tan / 10
        town = int(area_m2 // (one_tan * 10))
        remain_after_town = area_m2 - int(town * one_tan * 10)
        tan = int(remain_after_town // one_tan)
        remain_after_tan = remain_after_town - int(tan * one_tan)
        se = remain_after_tan / one_se
        return f"{town}町 {tan}反 {se:.1f}畝"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registered_totals = (
            RokunoheLandRegistry.objects.exclude(registered_land_category__isnull=True)
            .exclude(registered_land_category__exact="")
            .values("registered_land_category")
            .annotate(total_m2=Sum("registered_area"))
            .order_by("registered_land_category")
        )
        current_totals = (
            RokunoheLandRegistry.objects.exclude(current_land_category__isnull=True)
            .exclude(current_land_category__exact="")
            .values("current_land_category")
            .annotate(total_m2=Sum("current_area"))
            .order_by("current_land_category")
        )
        context["registered_totals"] = [
            {
                "category": row["registered_land_category"],
                "total_m2": row["total_m2"] or 0,
                "total_ha": (row["total_m2"] or 0) / 10000,
                "total_chobu": self._to_chobu(int(row["total_m2"] or 0)),
            }
            for row in registered_totals
        ]
        context["current_totals"] = [
            {
                "category": row["current_land_category"],
                "total_m2": row["total_m2"] or 0,
                "total_ha": (row["total_m2"] or 0) / 10000,
                "total_chobu": self._to_chobu(int(row["total_m2"] or 0)),
            }
            for row in current_totals
        ]
        return context
