import dataclasses
import os

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from soil_analysis.domain.service.chemical_import_service import ChemicalImportService
from soil_analysis.models import LandLedger


class Command(BaseCommand):
    """川田研究所フォーマットのExcelから化学分析データを取り込むDjangoコマンド

    川田研究所が提供する化学分析結果Excel(.xlsx)を読み込み、
    指定されたLandLedgerに紐づけてSoilChemicalMeasurementテーブルにデータを一括登録する。

    1つの圃場データを5ブロック（1,3,5,7,9）にコピーする仕様。

    使用方法:
        python manage.py chemical_load_data <excel_path> --land-ledger-id=<id>

    引数:
        excel_path: 川田研究所のExcelファイルパス（.xlsx）
        --land-ledger-id: 取り込み先のLandLedger ID（必須）
        既存データは常に更新されます（upsert）。
    """

    help = "Import chemical analysis data from Excel (Kawada format)"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to Excel file (.xlsx)")
        parser.add_argument(
            "--land-ledger-id",
            type=int,
            required=True,
            help="LandLedger ID to associate",
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]
        land_ledger_id = options["land_ledger_id"]

        if not os.path.exists(file_path):
            self.stderr.write(
                self.style.ERROR(f"エラー: ファイル '{file_path}' が見つかりません。")
            )
            return

        try:
            workbook = load_workbook(file_path, data_only=True)
        except PermissionError:
            self.stderr.write(
                self.style.ERROR(
                    f"エラー: ファイル '{file_path}' へのアクセスが拒否されました。\n"
                    f"Excel でファイルを開いている場合は、閉じてから再試行してください。"
                )
            )
            return
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"エラー: {str(e)}"))
            return

        try:
            # シート数が1であることを確認
            if len(workbook.sheetnames) != 1:
                self.stderr.write(
                    self.style.ERROR(
                        f"Excelファイルのシート数が1ではありません（{len(workbook.sheetnames)}枚）。"
                        f"シートは1枚にしてください。"
                    )
                )
                return
            worksheet = workbook.active
            parse_result = ChemicalImportService.parse_kawada_worksheet(worksheet)

            # パースエラーが存在する場合は処理を中断
            if parse_result.errors:
                for error in parse_result.errors:
                    self.stderr.write(self.style.ERROR(error))
                return

            # 取り込み対象行が存在しない場合は処理を中断
            if not parse_result.rows:
                self.stderr.write(self.style.WARNING("取り込み対象行がありません。"))
                return

            # LandLedgerの存在確認
            try:
                LandLedger.objects.get(id=land_ledger_id)
            except LandLedger.DoesNotExist:
                self.stderr.write(
                    self.style.ERROR(
                        f"LandLedger ID {land_ledger_id} が見つかりません。"
                    )
                )
                return

            # 保存用データ作成（コマンドの場合は指定された単一の帳簿に全行紐付ける）
            save_data = []
            for row in parse_result.rows:
                save_data.append(
                    {
                        "row_data": dataclasses.asdict(row),
                        "land_ledger_id": land_ledger_id,
                    }
                )

            result = ChemicalImportService.save_import_data(save_data)

            self.stdout.write(
                self.style.SUCCESS(
                    f"取り込み完了: 新規作成={result['created']}, 更新={result['updated']}, 警告=0"
                )
            )

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"エラー: {str(e)}"))
