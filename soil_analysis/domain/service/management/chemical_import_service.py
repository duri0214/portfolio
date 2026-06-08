import os
import re

from openpyxl import load_workbook

from soil_analysis.domain.repository.management.chemical_import_repository import (
    ChemicalImportRepository,
)
from soil_analysis.domain.valueobject.management.chemical_import_parser import (
    ChemicalImportParser,
)


class ChemicalImportService:
    """
    化学分析データのインポートプロセスの調整を行うサービス
    """

    @classmethod
    def import_from_excel(cls, file_path: str, land_ledger_id: int) -> dict:
        """
        Excelファイルから化学分析データを取り込み、指定された LandLedger に紐づけて保存する。
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"ファイル '{file_path}' が見つかりません。")

        try:
            workbook = load_workbook(file_path, data_only=True)
        except PermissionError:
            raise PermissionError(
                f"ファイル '{file_path}' へのアクセスが拒否されました。Excelで開いている場合は閉じてください。"
            )

        if len(workbook.sheetnames) != 1:
            raise ValueError(
                f"Excelファイルのシート数が1ではありません（{len(workbook.sheetnames)}枚）。シートは1枚にしてください。"
            )

        worksheet = workbook.active
        parse_result = ChemicalImportParser.parse_kawada_worksheet(worksheet)

        if parse_result.errors:
            for error_msg in parse_result.errors:
                row_number = None
                match = re.search(r"row=(\d+):", error_msg)
                if match:
                    row_number = int(match.group(1))

                ChemicalImportRepository.create_error(
                    row_number=row_number, land_name=None, message=error_msg
                )
            raise ValueError("\n".join(parse_result.errors))

        # 分析番号の重複チェック
        analysis_errors = ChemicalImportRepository.validate_analysis_numbers(
            parse_result.rows
        )
        if analysis_errors:
            for error_msg in analysis_errors:
                row_number = None
                match = re.search(r"row=(\d+):", error_msg)
                if match:
                    row_number = int(match.group(1))

                ChemicalImportRepository.create_error(
                    row_number=row_number, land_name=None, message=error_msg
                )
            raise ValueError("\n".join(analysis_errors))

        if not parse_result.rows:
            return {
                "created": 0,
                "updated": 0,
                "warning": "取り込み対象行がありません。",
            }

        if not ChemicalImportRepository.exists_ledger(land_ledger_id):
            raise ValueError(f"LandLedger ID {land_ledger_id} が見つかりません。")

        # 既存のロジックでは、Excel内の全行を一つの LandLedger ID に紐付けて保存している
        # ただし、同じ ledger_id に対して複数行ある場合は、ChemicalImportService.save_import_data では後のものを優先していた
        # ここでは単純に全行を処理する形にする（Repository側が適切にハンドルする）
        measurements_data = []
        for row in parse_result.rows:
            measurements_data.append(
                {"land_ledger_id": land_ledger_id, "record_values": row.to_dict()}
            )

        ChemicalImportRepository.delete_all_errors()
        result = ChemicalImportRepository.save_measurements(
            measurements_data, source_file=os.path.basename(file_path)
        )

        return result
